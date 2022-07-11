from flask_restful import  reqparse
import pandas as pd
import settings
import openpyxl

class ParserController:
    def __init__(self):
        self.reqparse = reqparse.RequestParser()
        self.excel_file = settings.EXCEL_FILE

    def read_excel(self):
        data_frames = pd.read_excel(self.excel_file)
        columns = list(data_frames)
        novels = data_frames.to_dict()[columns[1]]
        # to get novel name from hyperlink
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb['Sheet1']
        row_count = ws.max_row
        for row in range(3, row_count+1):
            hyperlink = ws.cell(row=row, column=2).value
            novel_name = hyperlink.split(',')[1]
            name = novel_name.split('"')[1]
            novels[row-2] = name
            data_frames.loc[data_frames.index[row-2],columns[1]] = name
        return data_frames, columns
    
    def parse_request(self):
        pass
    


